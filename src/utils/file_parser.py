"""
File parser for reading input variant files.
Supports CSV (comma, tab, semicolon delimited) and XLSX formats.
"""

import csv
import re
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from dataclasses import dataclass

from ..core.models import Variant, VariantType
from .logger import LoggerMixin


@dataclass
class ParseResult:
    """Result of parsing an input file."""
    variants: List[Variant]
    errors: List[str]
    warnings: List[str]
    total_rows: int
    successful_rows: int


class FileParser(LoggerMixin):
    """
    Parser for variant input files.
    Automatically detects file format and delimiter.
    """

    # Required column names (case-insensitive)
    GENE_COLUMNS = ['gen', 'gene', 'gene_symbol', 'gene symbol']
    TRANSCRIPT_COLUMNS = ['transkrypt', 'transcript', 'nm', 'refseq', 'transcript_id']
    VARIANT_COLUMNS = ['pozycja', 'position', 'pozycja danej zmiany', 'variant', 'hgvs', 'c.', 'hgvs_c']

    def __init__(self):
        """Initialize the file parser."""
        self._column_mapping: Dict[str, str] = {}

    def parse_file(self, file_path: str) -> ParseResult:
        """
        Parse a variant input file.

        Args:
            file_path: Path to the input file (CSV or XLSX)

        Returns:
            ParseResult containing parsed variants and any errors/warnings
        """
        path = Path(file_path)
        if not path.exists():
            return ParseResult(
                variants=[],
                errors=[f"File not found: {file_path}"],
                warnings=[],
                total_rows=0,
                successful_rows=0
            )

        extension = path.suffix.lower()
        self.log_info(f"Parsing file: {file_path} (format: {extension})")

        if extension == '.xlsx':
            return self._parse_xlsx(file_path)
        elif extension == '.csv':
            return self._parse_csv(file_path)
        else:
            return ParseResult(
                variants=[],
                errors=[f"Unsupported file format: {extension}. Use .csv or .xlsx"],
                warnings=[],
                total_rows=0,
                successful_rows=0
            )

    def _parse_xlsx(self, file_path: str) -> ParseResult:
        """Parse an XLSX file."""
        try:
            import openpyxl
        except ImportError:
            return ParseResult(
                variants=[],
                errors=["openpyxl library is required to read XLSX files. Install it with: pip install openpyxl"],
                warnings=[],
                total_rows=0,
                successful_rows=0
            )

        variants = []
        errors = []
        warnings = []

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active

            # Get headers from first row
            headers = []
            for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
                headers.append(str(cell.value).strip() if cell.value else "")

            # Map columns
            column_indices = self._map_columns(headers)
            if not column_indices:
                return ParseResult(
                    variants=[],
                    errors=["Could not identify required columns (gene, transcript, variant position)"],
                    warnings=[],
                    total_rows=0,
                    successful_rows=0
                )

            # Parse data rows
            row_number = 1
            total_rows = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_number += 1
                total_rows += 1

                # Skip empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                variant, row_errors, row_warnings = self._parse_row(
                    row, column_indices, row_number
                )

                if variant:
                    variants.append(variant)
                errors.extend(row_errors)
                warnings.extend(row_warnings)

            workbook.close()

        except Exception as e:
            self.log_exception(f"Error parsing XLSX file: {e}")
            errors.append(f"Error reading XLSX file: {str(e)}")

        return ParseResult(
            variants=variants,
            errors=errors,
            warnings=warnings,
            total_rows=total_rows,
            successful_rows=len(variants)
        )

    def _parse_csv(self, file_path: str) -> ParseResult:
        """Parse a CSV file with automatic delimiter detection."""
        variants = []
        errors = []
        warnings = []

        try:
            # Detect delimiter
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                sample = f.read(4096)
                f.seek(0)

            delimiter = self._detect_delimiter(sample)
            self.log_info(f"Detected delimiter: {repr(delimiter)}")

            with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.reader(f, delimiter=delimiter)

                # Get headers
                headers = next(reader)
                headers = [h.strip() for h in headers]

                # Map columns
                column_indices = self._map_columns(headers)
                if not column_indices:
                    return ParseResult(
                        variants=[],
                        errors=["Could not identify required columns (gene, transcript, variant position)"],
                        warnings=[],
                        total_rows=0,
                        successful_rows=0
                    )

                # Parse data rows
                row_number = 1
                total_rows = 0
                for row in reader:
                    row_number += 1
                    total_rows += 1

                    # Skip empty rows
                    if all((cell or '').strip() == '' for cell in row):
                        continue

                    variant, row_errors, row_warnings = self._parse_row(
                        row, column_indices, row_number
                    )

                    if variant:
                        variants.append(variant)
                    errors.extend(row_errors)
                    warnings.extend(row_warnings)

        except Exception as e:
            self.log_exception(f"Error parsing CSV file: {e}")
            errors.append(f"Error reading CSV file: {str(e)}")

        return ParseResult(
            variants=variants,
            errors=errors,
            warnings=warnings,
            total_rows=total_rows,
            successful_rows=len(variants)
        )

    def _detect_delimiter(self, sample: str) -> str:
        """Detect the delimiter used in CSV file."""
        # Count occurrences of common delimiters in first few lines
        delimiters = [',', '\t', ';']
        counts = {d: sample.count(d) for d in delimiters}

        # Use csv.Sniffer as fallback
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=''.join(delimiters))
            return dialect.delimiter
        except csv.Error:
            pass

        # Return delimiter with highest count
        return max(counts, key=counts.get)

    def _map_columns(self, headers: List[str]) -> Optional[Dict[str, int]]:
        """
        Map column headers to their indices.

        Returns:
            Dictionary mapping 'gene', 'transcript', 'variant' to column indices
        """
        mapping = {}
        headers_lower = [h.lower().strip() for h in headers]

        # Find gene column
        for col_name in self.GENE_COLUMNS:
            if col_name in headers_lower:
                mapping['gene'] = headers_lower.index(col_name)
                break

        # Find transcript column
        for col_name in self.TRANSCRIPT_COLUMNS:
            if col_name in headers_lower:
                mapping['transcript'] = headers_lower.index(col_name)
                break

        # Find variant column
        for col_name in self.VARIANT_COLUMNS:
            if col_name in headers_lower:
                mapping['variant'] = headers_lower.index(col_name)
                break

        # All three columns are required
        if len(mapping) == 3:
            self.log_info(f"Column mapping: {mapping}")
            return mapping

        # Try to find columns by partial match
        for i, header in enumerate(headers_lower):
            if 'gene' not in mapping and any(g in header for g in ['gen', 'gene']):
                mapping['gene'] = i
            elif 'transcript' not in mapping and any(t in header for t in ['transcript', 'nm_', 'refseq']):
                mapping['transcript'] = i
            elif 'variant' not in mapping and any(v in header for v in ['pozycja', 'position', 'variant', 'c.', 'hgvs']):
                mapping['variant'] = i

        if len(mapping) == 3:
            self.log_info(f"Column mapping (partial match): {mapping}")
            return mapping

        self.log_warning(f"Could not map all required columns. Found: {mapping}")
        return None

    def _parse_row(
        self,
        row: List[Any],
        column_indices: Dict[str, int],
        row_number: int
    ) -> Tuple[Optional[Variant], List[str], List[str]]:
        """
        Parse a single row from the input file.

        Returns:
            Tuple of (Variant or None, list of errors, list of warnings)
        """
        errors = []
        warnings = []

        try:
            # Check if row has enough columns
            max_idx = max(column_indices.values())
            if len(row) <= max_idx:
                errors.append(f"Row {row_number}: Row has {len(row)} columns, but column index {max_idx} is required")
                return None, errors, warnings

            # Extract values safely
            gene_idx = column_indices['gene']
            transcript_idx = column_indices['transcript']
            variant_idx = column_indices['variant']

            gene_val = row[gene_idx] if gene_idx < len(row) else None
            transcript_val = row[transcript_idx] if transcript_idx < len(row) else None
            variant_val = row[variant_idx] if variant_idx < len(row) else None

            gene = str(gene_val).strip() if gene_val else ""
            transcript = str(transcript_val).strip() if transcript_val else ""
            variant_str = str(variant_val).strip() if variant_val else ""

            # Basic validation
            if not gene:
                errors.append(f"Row {row_number}: Missing gene symbol")
                return None, errors, warnings

            if not transcript:
                errors.append(f"Row {row_number}: Missing transcript")
                return None, errors, warnings

            if not variant_str:
                errors.append(f"Row {row_number}: Missing variant position")
                return None, errors, warnings

            # Clean up transcript - extract NM_ if combined with variant
            transcript_clean, variant_clean = self._clean_transcript_variant(transcript, variant_str)

            # Parse variant type
            variant_type, cds_start, cds_end, ref_allele, alt_allele = self._parse_hgvs_basic(variant_clean)

            # Create variant object
            variant = Variant(
                row_number=row_number,
                gene_symbol=gene.upper(),
                transcript_accession=transcript_clean,
                hgvs_c=variant_clean,
                variant_type=variant_type,
                cds_start=cds_start,
                cds_end=cds_end,
                ref_allele=ref_allele,
                alt_allele=alt_allele
            )

            return variant, errors, warnings

        except IndexError:
            errors.append(f"Row {row_number}: Missing columns")
            return None, errors, warnings
        except Exception as e:
            errors.append(f"Row {row_number}: Error parsing - {str(e)}")
            return None, errors, warnings

    def _clean_transcript_variant(self, transcript: str, variant: str) -> Tuple[str, str]:
        """
        Clean and separate transcript and variant if they are combined.

        Args:
            transcript: Raw transcript string
            variant: Raw variant string

        Returns:
            Tuple of (cleaned transcript, cleaned variant)
        """
        # Sometimes transcript and variant are in the same field
        # e.g., "NM_002485.4:c.657_661delACAAA" or "NM_002485.4 c.657_661delACAAA"

        variant_clean = variant

        # Check if transcript contains variant notation
        if ':c.' in transcript or ' c.' in transcript:
            parts = re.split(r'[:\s](?=c\.)', transcript)
            if len(parts) >= 2:
                transcript = parts[0].strip()
                variant_clean = parts[1].strip()

        # Clean transcript - extract NM_xxxxx.x pattern
        nm_match = re.search(r'(NM_\d+)(?:\.(\d+))?', transcript, re.IGNORECASE)
        if nm_match:
            transcript = nm_match.group(0).upper()

        # Ensure variant starts with c.
        if not variant_clean.lower().startswith('c.'):
            variant_clean = 'c.' + variant_clean

        # Replace HTML entities (e.g., &gt; -> >)
        variant_clean = variant_clean.replace('&gt;', '>')
        variant_clean = variant_clean.replace('&lt;', '<')

        return transcript, variant_clean

    def _parse_hgvs_basic(
        self, hgvs: str
    ) -> Tuple[VariantType, int, int, str, str]:
        """
        Basic parsing of HGVS c. notation.
        Full validation is done later by HGVSValidator.

        Args:
            hgvs: HGVS c. notation string

        Returns:
            Tuple of (variant_type, cds_start, cds_end, ref_allele, alt_allele)
        """
        variant_type = VariantType.UNKNOWN
        cds_start = 0
        cds_end = 0
        ref_allele = ""
        alt_allele = ""

        # Remove 'c.' prefix
        if hgvs.lower().startswith('c.'):
            hgvs = hgvs[2:]

        # Try to match different variant patterns
        # Substitution: 657A>G
        sub_match = re.match(r'(\d+)([ACGT])>([ACGT])', hgvs, re.IGNORECASE)
        if sub_match:
            cds_start = cds_end = int(sub_match.group(1))
            ref_allele = sub_match.group(2).upper()
            alt_allele = sub_match.group(3).upper()
            variant_type = VariantType.SUBSTITUTION
            return variant_type, cds_start, cds_end, ref_allele, alt_allele

        # Deletion with range: 657_661delACAAA or 657_661del
        del_range_match = re.match(r'(\d+)_(\d+)del([ACGT]*)?', hgvs, re.IGNORECASE)
        if del_range_match:
            cds_start = int(del_range_match.group(1))
            cds_end = int(del_range_match.group(2))
            ref_allele = del_range_match.group(3).upper() if del_range_match.group(3) else ""
            variant_type = VariantType.DELETION
            return variant_type, cds_start, cds_end, ref_allele, alt_allele

        # Single position deletion: 657del or 657delA
        del_single_match = re.match(r'(\d+)del([ACGT])?', hgvs, re.IGNORECASE)
        if del_single_match:
            cds_start = cds_end = int(del_single_match.group(1))
            ref_allele = del_single_match.group(2).upper() if del_single_match.group(2) else ""
            variant_type = VariantType.DELETION
            return variant_type, cds_start, cds_end, ref_allele, alt_allele

        # Duplication with range: 657_661dup or 657_661dupACAAA
        dup_range_match = re.match(r'(\d+)_(\d+)dup([ACGT]*)?', hgvs, re.IGNORECASE)
        if dup_range_match:
            cds_start = int(dup_range_match.group(1))
            cds_end = int(dup_range_match.group(2))
            ref_allele = dup_range_match.group(3).upper() if dup_range_match.group(3) else ""
            variant_type = VariantType.DUPLICATION
            return variant_type, cds_start, cds_end, ref_allele, alt_allele

        # Single position duplication: 657dup or 657dupA
        dup_single_match = re.match(r'(\d+)dup([ACGT])?', hgvs, re.IGNORECASE)
        if dup_single_match:
            cds_start = cds_end = int(dup_single_match.group(1))
            ref_allele = dup_single_match.group(2).upper() if dup_single_match.group(2) else ""
            variant_type = VariantType.DUPLICATION
            return variant_type, cds_start, cds_end, ref_allele, alt_allele

        # Insertion: 657_658insACG
        ins_match = re.match(r'(\d+)_(\d+)ins([ACGT]+)', hgvs, re.IGNORECASE)
        if ins_match:
            cds_start = int(ins_match.group(1))
            cds_end = int(ins_match.group(2))
            alt_allele = ins_match.group(3).upper()
            variant_type = VariantType.INSERTION
            return variant_type, cds_start, cds_end, ref_allele, alt_allele

        # Delins: 657_661delACAAinsGT
        delins_match = re.match(r'(\d+)(?:_(\d+))?del([ACGT]*)ins([ACGT]+)', hgvs, re.IGNORECASE)
        if delins_match:
            cds_start = int(delins_match.group(1))
            cds_end = int(delins_match.group(2)) if delins_match.group(2) else cds_start
            ref_allele = delins_match.group(3).upper() if delins_match.group(3) else ""
            alt_allele = delins_match.group(4).upper()
            variant_type = VariantType.DELINS
            return variant_type, cds_start, cds_end, ref_allele, alt_allele

        # Try to extract at least position for unknown types
        pos_match = re.match(r'(\d+)', hgvs)
        if pos_match:
            cds_start = cds_end = int(pos_match.group(1))

        return variant_type, cds_start, cds_end, ref_allele, alt_allele
